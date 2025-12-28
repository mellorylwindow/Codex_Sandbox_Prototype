from __future__ import annotations

import csv
import hashlib
import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from html.parser import HTMLParser
from pathlib import Path
from typing import Dict, List, Optional, Tuple

# Inputs: your current intake location
IN_DIR = Path("notes/tax/intake/insurance/2025")
IN_DIR_BCBS = IN_DIR / "BCBSND"

OUT_DIR = Path("notes/tax/work/parsed/2025")
OUT_DIR.mkdir(parents=True, exist_ok=True)

OUT_NORM = OUT_DIR / "bcbsnd_claims.normalized.csv"
OUT_BAD  = OUT_DIR / "bcbsnd_claims.unparsed_rows.csv"
OUT_META = OUT_DIR / "bcbsnd_claims.import_log.txt"

FIELDNAMES = [
    "source_file",
    "source_type",              # xlsx|xls_html|xls_binary|csv
    "source_sha256",
    "row_id",

    "claim_id",
    "member",
    "provider",
    "service_date_from",
    "service_date_to",
    "received_date",
    "processed_date",
    "status",

    "billed_amount",
    "allowed_amount",
    "plan_paid",
    "patient_responsibility",
    "deductible",
    "copay",
    "coinsurance",

    "raw_category",
    "raw_note",
]

DATE_RX = re.compile(r"\b(\d{1,2})/(\d{1,2})/(\d{2,4})\b")
ISO_RX  = re.compile(r"^\d{4}-\d{2}-\d{2}$")
MONEY_RX = re.compile(r"[-]?\$?\s*([0-9]{1,3}(?:,[0-9]{3})*|[0-9]+)\.([0-9]{2})")

def sha256_file(p: Path) -> str:
    h = hashlib.sha256()
    with p.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()

def norm_date(s: str) -> str:
    if not s:
        return ""
    s = str(s).strip()
    if not s:
        return ""
    if ISO_RX.match(s):
        return s
    m = DATE_RX.search(s)
    if m:
        mm, dd, yy = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if yy < 100:
            yy += 2000
        try:
            return f"{yy:04d}-{mm:02d}-{dd:02d}"
        except Exception:
            return ""
    # try parse common datetime strings
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            dt = datetime.strptime(s[:10], fmt)
            return dt.strftime("%Y-%m-%d")
        except Exception:
            pass
    return ""

def norm_money(s) -> str:
    if s is None:
        return ""
    s = str(s).strip()
    if not s:
        return ""
    # allow plain numbers too
    try:
        # strip $ and commas
        cleaned = s.replace("$", "").replace(",", "")
        # sometimes parens mean negative
        neg = cleaned.startswith("(") and cleaned.endswith(")")
        if neg:
            cleaned = cleaned[1:-1]
        v = float(cleaned)
        if neg:
            v = -v
        return f"{v:.2f}"
    except Exception:
        pass

    m = MONEY_RX.search(s.replace(" ", ""))
    if not m:
        return ""
    whole = m.group(1).replace(",", "")
    cents = m.group(2)
    try:
        v = float(f"{whole}.{cents}")
        return f"{v:.2f}"
    except Exception:
        return ""

def canonical_header(h: str) -> str:
    h = (h or "").strip().lower()
    h = re.sub(r"\s+", " ", h)
    h = h.replace("\u00a0", " ")
    return h

# Broad mapping — we’ll refine after first run if needed.
COLMAP = {
    # identifiers
    "claim": "claim_id",
    "claim #": "claim_id",
    "claim number": "claim_id",
    "claim id": "claim_id",

    # person/provider
    "member": "member",
    "patient": "member",
    "subscriber": "member",
    "provider": "provider",
    "provider name": "provider",
    "facility": "provider",

    # dates
    "date of service": "service_date_from",
    "service date": "service_date_from",
    "from date": "service_date_from",
    "to date": "service_date_to",
    "received date": "received_date",
    "processed date": "processed_date",

    # status
    "status": "status",
    "claim status": "status",

    # amounts
    "billed amount": "billed_amount",
    "billed": "billed_amount",
    "charges": "billed_amount",

    "allowed amount": "allowed_amount",
    "allowed": "allowed_amount",

    "plan paid": "plan_paid",
    "paid by plan": "plan_paid",
    "insurer paid": "plan_paid",

    "patient responsibility": "patient_responsibility",
    "your responsibility": "patient_responsibility",
    "member responsibility": "patient_responsibility",

    "deductible": "deductible",
    "copay": "copay",
    "co-pay": "copay",
    "coinsurance": "coinsurance",
    "co-insurance": "coinsurance",

    # extras
    "category": "raw_category",
    "note": "raw_note",
    "notes": "raw_note",
    "description": "raw_note",
}

def map_row(row: Dict[str, str]) -> Dict[str, str]:
    out = {k: "" for k in FIELDNAMES}
    # Map columns
    for k, v in row.items():
        ck = canonical_header(k)
        if ck in COLMAP:
            out[COLMAP[ck]] = (v or "").strip()

    # Normalize fields
    out["service_date_from"] = norm_date(out["service_date_from"])
    out["service_date_to"]   = norm_date(out["service_date_to"])
    out["received_date"]     = norm_date(out["received_date"])
    out["processed_date"]    = norm_date(out["processed_date"])

    for f in ("billed_amount","allowed_amount","plan_paid","patient_responsibility","deductible","copay","coinsurance"):
        out[f] = norm_money(out[f])

    # Some exports use one DOS range column like "01/01/2025 - 01/02/2025"
    if out["service_date_from"] and not out["service_date_to"]:
        # ok
        pass
    if not out["service_date_from"]:
        # try to locate any date in raw_note
        out["service_date_from"] = norm_date(out.get("raw_note",""))

    return out

# -------- HTML table parser (many “.xls” exports are actually HTML) --------
class TableExtractor(HTMLParser):
    def __init__(self):
        super().__init__()
        self.in_td = False
        self.in_th = False
        self.in_tr = False
        self.current_cell: List[str] = []
        self.current_row: List[str] = []
        self.rows: List[List[str]] = []
        self.tables: List[List[List[str]]] = []
        self.in_table = False

    def handle_starttag(self, tag, attrs):
        tag = tag.lower()
        if tag == "table":
            self.in_table = True
            self.rows = []
        elif tag == "tr" and self.in_table:
            self.in_tr = True
            self.current_row = []
        elif tag in ("td","th") and self.in_tr:
            self.in_td = (tag == "td")
            self.in_th = (tag == "th")
            self.current_cell = []

    def handle_endtag(self, tag):
        tag = tag.lower()
        if tag in ("td","th") and self.in_tr:
            txt = "".join(self.current_cell).strip()
            txt = re.sub(r"\s+", " ", txt)
            self.current_row.append(txt)
            self.in_td = False
            self.in_th = False
            self.current_cell = []
        elif tag == "tr" and self.in_table:
            self.in_tr = False
            if any(c.strip() for c in self.current_row):
                self.rows.append(self.current_row)
            self.current_row = []
        elif tag == "table":
            self.in_table = False
            if self.rows:
                self.tables.append(self.rows)
            self.rows = []

    def handle_data(self, data):
        if (self.in_td or self.in_th) and self.in_tr:
            self.current_cell.append(data)

def read_xlsx_rows(p: Path) -> Tuple[List[Dict[str,str]], str]:
    try:
        from openpyxl import load_workbook
    except Exception as e:
        raise RuntimeError("openpyxl missing. Install with: pip install openpyxl") from e

    wb = load_workbook(p, data_only=True)
    # try each sheet; pick the one with the most rows
    best = None
    best_rows = 0
    for name in wb.sheetnames:
        ws = wb[name]
        values = list(ws.values)
        if len(values) > best_rows:
            best_rows = len(values)
            best = values

    if not best or best_rows < 2:
        return [], "xlsx_empty"

    # find header row: the first row that contains a claim-ish column
    header_idx = None
    for i, row in enumerate(best[:30]):
        if not row:
            continue
        joined = " ".join(str(x or "") for x in row).lower()
        if "claim" in joined and ("service" in joined or "date" in joined or "provider" in joined):
            header_idx = i
            break
    if header_idx is None:
        header_idx = 0

    headers = [str(h or "").strip() for h in best[header_idx]]
    out: List[Dict[str,str]] = []
    for row in best[header_idx+1:]:
        if not row:
            continue
        d = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            v = row[i] if i < len(row) else ""
            d[h] = "" if v is None else str(v)
        # skip totally blank
        if not any(str(v).strip() for v in d.values()):
            continue
        out.append(d)
    return out, "xlsx"

def read_xls_rows(p: Path) -> Tuple[List[Dict[str,str]], str]:
    # Many “.xls” exports are HTML.
    raw = p.read_bytes()
    # try decode as text
    text = None
    for enc in ("utf-8", "utf-8-sig", "cp1252", "latin1"):
        try:
            text = raw.decode(enc)
            break
        except Exception:
            continue
    if text is None:
        return [], "xls_binary"

    if "<table" in text.lower():
        parser = TableExtractor()
        parser.feed(text)
        if not parser.tables:
            return [], "xls_html_no_tables"

        # choose largest table
        tables = sorted(parser.tables, key=lambda t: (len(t), max((len(r) for r in t), default=0)), reverse=True)
        tab = tables[0]
        # first row = header
        headers = tab[0]
        rows = tab[1:]
        out: List[Dict[str,str]] = []
        for r in rows:
            d = {}
            for i, h in enumerate(headers):
                h = str(h or "").strip()
                if not h:
                    continue
                d[h] = str(r[i] if i < len(r) else "").strip()
            if any(v.strip() for v in d.values()):
                out.append(d)
        return out, "xls_html"

    # fallback: tab-delimited
    if "\t" in text and "\n" in text:
        lines = [ln.strip("\r") for ln in text.splitlines() if ln.strip()]
        if len(lines) < 2:
            return [], "xls_text_too_small"
        headers = [h.strip() for h in lines[0].split("\t")]
        out = []
        for ln in lines[1:]:
            parts = [p.strip() for p in ln.split("\t")]
            d = {headers[i]: (parts[i] if i < len(parts) else "") for i in range(len(headers))}
            if any(v.strip() for v in d.values()):
                out.append(d)
        return out, "xls_tsv"

    return [], "xls_unrecognized"

def write_csv(path: Path, rows: List[Dict[str,str]]):
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=FIELDNAMES)
        w.writeheader()
        for r in rows:
            w.writerow({k: r.get(k, "") for k in FIELDNAMES})

def main():
    sources = []
    if IN_DIR.exists():
        sources += list(IN_DIR.glob("*.xls")) + list(IN_DIR.glob("*.xlsx")) + list(IN_DIR.glob("*.csv"))
    if IN_DIR_BCBS.exists():
        sources += list(IN_DIR_BCBS.glob("*.xls")) + list(IN_DIR_BCBS.glob("*.xlsx")) + list(IN_DIR_BCBS.glob("*.csv"))

    sources = sorted({p.resolve() for p in sources if p.is_file()})

    log_lines = []
    norm_rows: List[Dict[str,str]] = []
    bad_rows: List[Dict[str,str]] = []

    if not sources:
        print(f"no BCBSND claim exports found in:")
        print(f"  {IN_DIR}")
        print(f"  {IN_DIR_BCBS}")
        return

    for p in sources:
        sha = sha256_file(p)
        suffix = p.suffix.lower()
        src_type = "<?>"
        raw_rows: List[Dict[str,str]] = []

        try:
            if suffix == ".xlsx":
                raw_rows, src_type = read_xlsx_rows(p)
            elif suffix == ".xls":
                raw_rows, src_type = read_xls_rows(p)
            elif suffix == ".csv":
                src_type = "csv"
                with p.open(newline="", encoding="utf-8-sig") as f:
                    r = csv.DictReader(f)
                    raw_rows = list(r)
            else:
                continue
        except Exception as e:
            log_lines.append(f"[FAIL] {p.name} ({suffix}): {e}")
            continue

        log_lines.append(f"[OK] {p.name} type={src_type} rows={len(raw_rows)}")

        for i, rr in enumerate(raw_rows, start=1):
            mapped = map_row(rr)

            mapped["source_file"] = str(p).replace("\\","/")
            mapped["source_type"] = src_type
            mapped["source_sha256"] = sha
            mapped["row_id"] = f"{p.name}::{i}"

            # minimal validity checks
            has_any_money = any(mapped.get(k) for k in ("patient_responsibility","plan_paid","allowed_amount","billed_amount"))
            has_date = bool(mapped.get("service_date_from"))

            # if claim_id empty, try to locate from any column that contains "CLM"
            if not mapped.get("claim_id"):
                for k, v in rr.items():
                    if "claim" in canonical_header(k) and str(v).strip():
                        mapped["claim_id"] = str(v).strip()
                        break

            if not has_date or not has_any_money:
                bad_rows.append(mapped)
            else:
                norm_rows.append(mapped)

    write_csv(OUT_NORM, norm_rows)
    write_csv(OUT_BAD, bad_rows)

    # Summaries (by month of service_date_from)
    month_tot = defaultdict(float)
    month_cnt = defaultdict(int)

    def f2(x: str) -> float:
        try:
            return float(x)
        except Exception:
            return 0.0

    for r in norm_rows:
        m = (r.get("service_date_from") or "")[:7]
        if not m:
            continue
        month_cnt[m] += 1
        month_tot[(m,"patient_responsibility")] += f2(r.get("patient_responsibility",""))
        month_tot[(m,"plan_paid")] += f2(r.get("plan_paid",""))
        month_tot[(m,"deductible")] += f2(r.get("deductible",""))
        month_tot[(m,"copay")] += f2(r.get("copay",""))
        month_tot[(m,"coinsurance")] += f2(r.get("coinsurance",""))

    out_summary = OUT_DIR / "bcbsnd_claims.summary_by_month.csv"
    with out_summary.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["month","tx_count","patient_responsibility","plan_paid","deductible","copay","coinsurance"])
        for m in sorted(month_cnt.keys()):
            w.writerow([
                m,
                month_cnt[m],
                f"{month_tot[(m,'patient_responsibility')]:.2f}",
                f"{month_tot[(m,'plan_paid')]:.2f}",
                f"{month_tot[(m,'deductible')]:.2f}",
                f"{month_tot[(m,'copay')]:.2f}",
                f"{month_tot[(m,'coinsurance')]:.2f}",
            ])

    OUT_META.write_text("\n".join(log_lines) + "\n", encoding="utf-8")

    # quick console stats
    dates = [r["service_date_from"] for r in norm_rows if r.get("service_date_from")]
    print("wrote:")
    print(f"  {OUT_NORM} rows: {len(norm_rows)}")
    print(f"  {OUT_BAD} rows: {len(bad_rows)}")
    print(f"  {out_summary} rows: {len(month_cnt)}")
    if dates:
        print(f"service date coverage: {min(dates)} -> {max(dates)}")
    print(f"log: {OUT_META}")

if __name__ == "__main__":
    main()
