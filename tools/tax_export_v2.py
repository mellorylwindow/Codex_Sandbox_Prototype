#!/usr/bin/env python3
"""
tax_export_v2.py — clean rewrite (v2.0)

Purpose
-------
Read a batch manifest (jsonl) + OCR text outputs (sha256.txt) and export:
  - CSV:  one row per canonical document (or all docs)
  - XLSX: Lines + Summary + ByVendor + NeedsReview

Key points
----------
- Manifest rows may include: sha256, original_name, status (ingested|duplicate), src_rel, dest_rel
- OCR text files live in: --text-dir and are named <sha256>.txt (full hash). Prefix fallback supported.
- --only-canonical filters out manifest rows where status == "duplicate"
- Vendor extraction is resilient against OCR junk (addresses, card lines, signoffs, emails, etc.)
- Amount extraction prefers labeled fields (Amount Due / Total Due / Balance Due / Patient Responsibility)
- Date extraction prefers service/visit dates in the selected --tax-year when possible

Example
-------
python tools/tax_export_v2.py \
  --manifest "out/tax/batches/2025-12-25_134023/manifest.jsonl" \
  --text-dir "tax_intake/30_extracted/text" \
  --only-canonical \
  --tax-year 2025 \
  --out-xlsx "tax_intake/40_reports/tax_lines_2025-12-25_134023_v2.xlsx" \
  --out-csv  "tax_intake/40_reports/tax_lines_2025-12-25_134023_v2.csv" \
  --debug
"""

from __future__ import annotations

import argparse
import csv
import json
import re
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import pandas as pd


# -----------------------------
# Regex building blocks
# -----------------------------

MONEY_RE = re.compile(r"(?<!\w)\$?\s*([0-9]{1,3}(?:,[0-9]{3})*|[0-9]+)\.([0-9]{2})(?!\w)")
MMDDYY_RE = re.compile(r"(?<!\d)(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})(?!\d)")
ZIP_RE = re.compile(r"\b\d{5}(?:-\d{4})?\b")
PHONE_RE = re.compile(r"\b(?:\(?\d{3}\)?[-.\s]?)\d{3}[-.\s]?\d{4}\b")
EMAIL_RE = re.compile(r"\b[a-z0-9._%+\-]+@[a-z0-9.\-]+\.[a-z]{2,}\b", re.I)

MONTHS = {
    "jan": 1, "january": 1,
    "feb": 2, "february": 2,
    "mar": 3, "march": 3,
    "apr": 4, "april": 4,
    "may": 5,
    "jun": 6, "june": 6,
    "jul": 7, "july": 7,
    "aug": 8, "august": 8,
    "sep": 9, "sept": 9, "september": 9,
    "oct": 10, "october": 10,
    "nov": 11, "november": 11,
    "dec": 12, "december": 12,
}

MONTHNAME_DATE_RE = re.compile(
    r"\b("
    r"jan(?:uary)?|feb(?:ruary)?|mar(?:ch)?|apr(?:il)?|may|"
    r"jun(?:e)?|jul(?:y)?|aug(?:ust)?|sep(?:t)?(?:ember)?|"
    r"oct(?:ober)?|nov(?:ember)?|dec(?:ember)?"
    r")\s+(\d{1,2})(?:st|nd|rd|th)?\,?\s+(\d{4})\b",
    re.I,
)

CARD_RE = re.compile(
    r"\b(credit|debit)\s*card\b|\bcard\s*number\b|\bvisa\b|\bmaster\s*card\b|\bmastercard\b|\bamex\b|\bdiscover\b",
    re.I,
)


def norm_space(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "")).strip()


def to_float_money(m: re.Match) -> Optional[float]:
    try:
        whole = m.group(1).replace(",", "")
        cents = m.group(2)
        return float(f"{whole}.{cents}")
    except Exception:
        return None


def find_all_money(text: str) -> List[float]:
    vals: List[float] = []
    for m in MONEY_RE.finditer(text or ""):
        v = to_float_money(m)
        if v is not None:
            vals.append(v)
    return vals


def parse_mmddyy(m: re.Match) -> Optional[date]:
    try:
        mm = int(m.group(1))
        dd = int(m.group(2))
        yy = int(m.group(3))
        if yy < 100:
            yy += 2000
        return date(yy, mm, dd)
    except Exception:
        return None


def parse_monthname_date(m: re.Match) -> Optional[date]:
    try:
        mon = m.group(1).lower()
        mon = mon[:3] if mon not in MONTHS else mon
        mm = MONTHS.get(mon.lower(), MONTHS.get(m.group(1).lower(), 0))
        dd = int(m.group(2))
        yy = int(m.group(3))
        if not mm:
            return None
        return date(yy, mm, dd)
    except Exception:
        return None


def all_dates(text: str) -> List[date]:
    out: List[date] = []
    for m in MMDDYY_RE.finditer(text or ""):
        d = parse_mmddyy(m)
        if d:
            out.append(d)
    for m in MONTHNAME_DATE_RE.finditer(text or ""):
        d = parse_monthname_date(m)
        if d:
            out.append(d)
    return out


def redact_text_for_export(text: str) -> str:
    """Light redaction for snippets (keep useful content; reduce risk)."""
    t = text or ""
    t = EMAIL_RE.sub("[EMAIL REDACTED]", t)
    t = re.sub(r"\bRX\s*#?\s*[:\-]?\s*[A-Z0-9\-]{4,}\b", "RX [REDACTED]", t, flags=re.I)
    t = re.sub(r"\b(i-?code)\b.*", "I-CODE [REDACTED]", t, flags=re.I)
    t = re.sub(r"\b(account|reference)\s*(number)?\s*[:\-]?\s*[0-9]{6,}\b", r"\1 [REDACTED]", t, flags=re.I)
    return t


# -----------------------------
# Categorization rules
# -----------------------------

# (regex, (category, subcategory))
CATEGORY_RULES: List[Tuple[re.Pattern, Tuple[str, str]]] = [
    # Medical (Schedule A)
    (re.compile(r"\b(community\s+radiology|radiology\s+associates|cra\b|payradne\w{0,4}bill\.com)\b", re.I),
     ("Medical", "Imaging / Radiology")),
    (re.compile(r"\b(cblpath|patholog|lab\s+corp|quest\s+diagnostic|diagnostic\s+lab)\b", re.I),
     ("Medical", "Lab / Pathology")),
    (re.compile(r"\b(privia|athenahealth|patient\s+responsibility|office\s+visit|visit\s+on)\b", re.I),
     ("Medical", "Provider / Office Visit")),
    (re.compile(r"\b(cvs\s+pharmacy|prescription\s+information|rx\s+information|date\s+filled|promised)\b", re.I),
     ("Medical", "Prescriptions / Pharmacy")),

    # Taxes (Schedule A SALT if itemizing)
    (re.compile(r"\b(county\s+of\s+fairfax|personal\s+property\s+tax|tax\s+bill|vehicle\s+assessed)\b", re.I),
     ("Taxes", "Personal Property Tax")),

    # Other (track, not necessarily deductible)
    (re.compile(r"\b(receivables\s+management\s+systems|this\s+is\s+an\s+attempt\s+to\s+collect\s+a\s+debt|debt\s+collector)\b", re.I),
     ("Other", "Collections / Admin")),
]

DEFAULT_CATEGORY = ("Uncategorized", "Review")


def classify_category(text: str, vendor: str) -> Tuple[str, str]:
    hay = f"{vendor}\n{text}"
    for rx, cat in CATEGORY_RULES:
        if rx.search(hay):
            return cat
    return DEFAULT_CATEGORY


def likely_deductible_bucket(cat: str) -> bool:
    # Conservative: only mark obvious buckets as "likely"
    return cat in {"Medical", "Taxes", "Charity", "Interest"}


# -----------------------------
# Vendor detection
# -----------------------------

SIGNOFF_REJECT = re.compile(
    r"\b(sincerely|regards|thank\s+you|your\s+receipt|receipt\s+email|please\s+retain|"
    r"privilege\s+to\s+serve|customer\s+service)\b",
    re.I
)

IDENTITY_BLACKLIST = {
    "member id",
    "policy",
    "group #",
    "group no",
    "patient:",
    "subscriber",
    "statement of benefits",
    "explanation of benefits",
    "claim number",
    "receipt email",
    "email",
    "credit card",
    "card number",
}

KNOWN_VENDOR_RULES: List[Tuple[re.Pattern, str]] = [
    # CRA portal domain gets OCR-mangled (net/nelt/neit etc) — be tolerant.
    (re.compile(r"\b(?:cra\.)?payradne\w{0,4}bill\.com\b", re.I), "Community Radiology Associates"),
    (re.compile(r"\bcommunity\s+radiology\b|\bradiology\s+associates\b|\bcra\b", re.I), "Community Radiology Associates"),
    (re.compile(r"\bcblpath\b", re.I), "CBLPath, Inc."),
    (re.compile(r"\bprivia\b|\bathenahealth\b", re.I), "Privia / Athenahealth"),
    (re.compile(r"\bcvs\b.*\bpharmacy\b|\bcvs\s+pharmacy\b", re.I), "CVS Pharmacy"),
    (re.compile(r"\bcounty\s+of\s+fairfax\b", re.I), "County of Fairfax (VA)"),
    (re.compile(r"\breceivables\s+management\s+systems\b|\brmscollect\b", re.I), "Receivables Management Systems"),
    (re.compile(r"\btarget\b", re.I), "Target"),
]


def looks_like_addressish(line: str) -> bool:
    l = (line or "").lower().strip()
    if not l:
        return True

    # Numeric street start (catches "2362 Antiqua C1" OCR variants)
    if re.match(r"^\d{1,6}\s+[a-z]", l):
        return True

    if "p.o." in l or "po box" in l:
        return True
    if "suite" in l or " ste " in f" {l} " or " apt " in f" {l} ":
        return True
    if ZIP_RE.search(line):
        return True
    if PHONE_RE.search(line):
        return True
    if re.search(r"\b(street|st\b|road|rd\b|avenue|ave\b|blvd|boulevard|lane|ln\b|drive|dr\b|court|ct\b|circle|cir\b)\b", l):
        return True
    return False


def looks_like_identity_line(line: str) -> bool:
    s = norm_space(line)
    sl = s.lower()
    if EMAIL_RE.search(s):
        return True
    if CARD_RE.search(s):
        return True
    for tok in IDENTITY_BLACKLIST:
        if tok in sl:
            return True
    return False


def vendor_score(line: str) -> int:
    s = norm_space(line)
    sl = s.lower()

    if not s:
        return -10_000
    if looks_like_addressish(s):
        return -9_500
    if looks_like_identity_line(s):
        return -9_300
    if SIGNOFF_REJECT.search(s):
        return -9_200
    if CARD_RE.search(s):
        return -9_400

    # Reject "table headers" / common boilerplate
    if re.search(r"\b(service\s+date|description|units|charge|total|deductible|coinsurance|copay)\b", sl):
        return -6_000

    score = 0

    # Prefer shorter but not tiny lines
    n = len(s)
    if 4 <= n <= 55:
        score += 500
    elif 56 <= n <= 90:
        score += 150
    else:
        score -= 300

    # Prefer letter-heavy lines
    letters = sum(ch.isalpha() for ch in s)
    digits = sum(ch.isdigit() for ch in s)
    if letters >= 6:
        score += 200
    if digits >= 4:
        score -= 150

    # Prefer org-ish tokens
    if re.search(r"\b(inc|llc|associates|association|medical|clinic|pharmacy|radiology|county|hospital)\b", sl):
        score += 600

    # Penalize if it looks like a sentence
    if re.search(r"\b(the|and|from|to|for)\b", sl) and n > 40:
        score -= 200

    return score


def detect_vendor(text: str) -> str:
    t = text or ""
    # First: known vendor patterns anywhere in text
    for rx, name in KNOWN_VENDOR_RULES:
        if rx.search(t):
            return name

    # Second: score candidate lines near the top
    lines = [norm_space(x) for x in t.splitlines()]
    lines = [x for x in lines if x]

    # Some docs put vendor a bit lower; consider first 35 non-empty lines
    candidates = lines[:35] if lines else []
    if not candidates:
        return "Unknown"

    best = max(candidates, key=vendor_score)
    if vendor_score(best) < 0:
        return "Unknown"
    return best


# -----------------------------
# Amount + status extraction
# -----------------------------

@dataclass
class Amounts:
    due: Optional[float]
    paid: Optional[float]
    status: str  # paid | estimate | due | unknown


LABEL_MONEY_PATTERNS: List[Tuple[str, re.Pattern]] = [
    ("due", re.compile(r"\b(total\s+due|amount\s+due|balance\s+due|due\s+now|patient\s+responsibility)\b\D{0,40}(\$?\s*[0-9,]+\.[0-9]{2})", re.I)),
    ("paid", re.compile(r"\b(amount\s+paid|payment\s+received|you\s+paid|paid\s+amount)\b\D{0,40}(\$?\s*[0-9,]+\.[0-9]{2})", re.I)),
]


def parse_amounts(text: str) -> Amounts:
    tl = (text or "").lower()

    # Status signals
    is_estimate = bool(re.search(r"\b(estimate|estimated|this\s+is\s+not\s+a\s+bill)\b", tl))
    has_paid_signal = bool(re.search(r"\b(payment\s+received|paid\s+in\s+full|paid)\b", tl))
    has_due_signal = bool(re.search(r"\b(amount\s+due|balance\s+due|total\s+due|due\s+now)\b", tl))

    due: Optional[float] = None
    paid: Optional[float] = None

    # Labeled captures (best)
    for kind, rx in LABEL_MONEY_PATTERNS:
        for m in rx.finditer(text or ""):
            raw = m.group(2) if kind == "due" else m.group(2)
            raw = raw.replace("$", "").replace(",", "").strip()
            try:
                val = float(raw)
            except Exception:
                continue
            if kind == "due" and due is None:
                due = val
            if kind == "paid" and paid is None:
                paid = val

    # If labeled due not found, fallback to "reasonable" money heuristic:
    # - Prefer smaller values for pharmacy
    # - Prefer the largest value under $50k for bills
    if due is None:
        vals = [v for v in find_all_money(text) if 0 <= v <= 50_000]
        if vals:
            # If looks like pharmacy, smaller values tend to be "amount due"
            if re.search(r"\b(cvs|pharmacy|rx|date\s+filled|promised)\b", tl):
                small = [v for v in vals if v <= 150]
                due = small[-1] if small else vals[-1]
            else:
                due = max(vals)

    # If due is explicitly $0.00 and we have paid signals, mark paid
    if due is not None and abs(due) < 1e-9 and has_paid_signal:
        status = "paid"
    elif is_estimate:
        status = "estimate"
    elif has_due_signal or (due is not None and due > 0):
        status = "due"
    elif has_paid_signal:
        status = "paid"
    else:
        status = "unknown"

    return Amounts(due=due, paid=paid, status=status)


# -----------------------------
# Date extraction
# -----------------------------

DATE_LABELS = {
    "service_date": ("service date", "visit", "date of service", "dos"),
    "statement_date": ("statement date", "statement", "bill date", "issued"),
    "due_date": ("due date", "pay by", "payment due", "due by"),
}


def pick_best_date(text: str, tax_year: int, prefer_labels: Iterable[str]) -> Optional[date]:
    """Pick best date: prefer near label windows, and prefer within tax_year."""
    t = text or ""
    tl = t.lower()

    scored: List[Tuple[int, date]] = []

    # First: near label windows
    for lab in prefer_labels:
        idx = tl.find(lab.lower())
        if idx >= 0:
            window = t[max(0, idx - 40): idx + 180]
            for d in all_dates(window):
                score = 0
                if d.year != tax_year:
                    score += 50
                # Slightly prefer newer dates within the year
                score += (tax_year - d.year) * 5
                scored.append((score, d))

    # Second: global dates
    for d in all_dates(t):
        score = 20
        if d.year != tax_year:
            score += 50
        scored.append((score, d))

    if not scored:
        return None

    # Sort by score then newest date
    scored.sort(key=lambda x: (x[0], -int(x[1].strftime("%Y%m%d"))))
    return scored[0][1]


def parse_dates(text: str, tax_year: int) -> Tuple[Optional[date], Optional[date], Optional[date]]:
    service = pick_best_date(text, tax_year, DATE_LABELS["service_date"])
    statement = pick_best_date(text, tax_year, DATE_LABELS["statement_date"])
    due = pick_best_date(text, tax_year, DATE_LABELS["due_date"])

    # If service missing but statement exists, use statement as service-ish for sorting
    if service is None and statement is not None:
        service = statement

    return service, statement, due


# -----------------------------
# Manifest + OCR text loading
# -----------------------------

def read_manifest_jsonl(path: Path) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            rows.append(json.loads(line))
    return rows


def is_duplicate_row(r: Dict[str, Any]) -> bool:
    return str(r.get("status", "")).lower() == "duplicate"


def collapse_to_canonical(rows: List[Dict[str, Any]], debug: bool = False) -> List[Dict[str, Any]]:
    """If manifest has duplicates, canonical are status != duplicate."""
    before = len(rows)
    out = [r for r in rows if not is_duplicate_row(r)]
    if debug:
        print(f"[debug] canonical collapse: {before} -> {len(out)} (dropped {before - len(out)} dups)")
    return out


def load_ocr_text(text_dir: Path, sha256: str) -> Optional[str]:
    """Load OCR text by full hash. If not found, try prefix match."""
    if not sha256:
        return None

    p_full = text_dir / f"{sha256}.txt"
    if p_full.exists():
        return p_full.read_text(encoding="utf-8", errors="replace")

    # Prefix fallback (rare)
    pref = sha256[:12]
    hits = list(text_dir.glob(f"{pref}*.txt"))
    if hits:
        return hits[0].read_text(encoding="utf-8", errors="replace")

    return None


def source_rel(r: Dict[str, Any]) -> str:
    # Prefer dest_rel if present, else src_rel, else empty
    return str(r.get("dest_rel") or r.get("src_rel") or "")


# -----------------------------
# Export rows
# -----------------------------

def build_rows(
    manifest_rows: List[Dict[str, Any]],
    text_dir: Path,
    tax_year: int,
    only_canonical: bool,
    debug: bool,
) -> Tuple[List[Dict[str, Any]], int]:
    rows = manifest_rows
    if only_canonical:
        rows = collapse_to_canonical(rows, debug=debug)

    out: List[Dict[str, Any]] = []
    missing_text = 0

    for i, r in enumerate(rows, 1):
        sha = str(r.get("sha256", ""))
        orig = str(r.get("original_name", ""))
        status = str(r.get("status", "")).lower() or "unknown"
        rel = source_rel(r)

        text = load_ocr_text(text_dir, sha)
        if not text:
            missing_text += 1
            text = ""

        vendor = detect_vendor(text)
        cat, subcat = classify_category(text, vendor)
        svc, stmt, due_dt = parse_dates(text, tax_year)
        amts = parse_amounts(text)

        snippet = redact_text_for_export(text)
        snippet = norm_space(snippet)[:220]

        row = {
            "tax_year": tax_year,
            "sha256": sha,
            "sha12": sha[:12],
            "original_name": orig,
            "manifest_status": status,  # ingested|duplicate|unknown
            "source_rel": rel,
            "vendor": vendor,
            "category": cat,
            "subcategory": subcat,
            "doc_status": amts.status,  # paid|estimate|due|unknown
            "service_date": svc.isoformat() if svc else "",
            "statement_date": stmt.isoformat() if stmt else "",
            "due_date": due_dt.isoformat() if due_dt else "",
            "expense_due": amts.due,
            "expense_paid": amts.paid,
            "likely_deductible": bool(likely_deductible_bucket(cat)),
            "notes": "",
            "snippet": snippet,
        }

        out.append(row)

        if debug and i <= 10:
            print(
                f"[debug] row {i}: {row['sha12']} vendor={row['vendor']!r} "
                f"svc={row['service_date']!r} paid={row['expense_paid']} due={row['expense_due']} "
                f"cat={row['category']}/{row['subcategory']} status={row['doc_status']}"
            )

    return out, missing_text


def write_csv(rows: List[Dict[str, Any]], out_csv: Path) -> None:
    out_csv.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        out_csv.write_text("", encoding="utf-8")
        return

    cols = list(rows[0].keys())
    with out_csv.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=cols)
        w.writeheader()
        for r in rows:
            w.writerow(r)


def autosize_columns(ws) -> None:
    """Small convenience: set column widths based on content lengths."""
    # ws is an openpyxl worksheet
    try:
        from openpyxl.utils import get_column_letter  # type: ignore
    except Exception:
        return

    max_width: Dict[int, int] = {}
    for row in ws.iter_rows(values_only=True):
        for j, v in enumerate(row, 1):
            if v is None:
                continue
            s = str(v)
            max_width[j] = max(max_width.get(j, 0), min(len(s), 60))
    for j, w in max_width.items():
        ws.column_dimensions[get_column_letter(j)].width = max(10, min(w + 2, 60))


def build_summaries(df: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    # Ensure numeric
    df = df.copy()
    df["expense_due"] = pd.to_numeric(df["expense_due"], errors="coerce")
    df["expense_paid"] = pd.to_numeric(df["expense_paid"], errors="coerce")

    # Summary by category/subcategory
    cat_sum = (
        df.groupby(["category", "subcategory"], dropna=False)
          .agg(
              docs=("sha256", "count"),
              total_due=("expense_due", "sum"),
              total_paid=("expense_paid", "sum"),
          )
          .reset_index()
          .sort_values(["category", "subcategory"])
    )

    # By vendor
    vend_sum = (
        df.groupby(["vendor", "category"], dropna=False)
          .agg(
              docs=("sha256", "count"),
              total_due=("expense_due", "sum"),
              total_paid=("expense_paid", "sum"),
          )
          .reset_index()
          .sort_values(["total_due", "docs"], ascending=[False, False])
    )

    # Needs review
    needs_review = df[
        (df["category"] == "Uncategorized")
        | (df["doc_status"] == "unknown")
        | (df["vendor"] == "Unknown")
    ].copy()

    return cat_sum, vend_sum, needs_review


def write_xlsx(df: pd.DataFrame, out_xlsx: Path, debug: bool = False) -> None:
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)

    cat_sum, vend_sum, needs_review = build_summaries(df)

    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as xw:
        df.to_excel(xw, index=False, sheet_name="Lines")
        cat_sum.to_excel(xw, index=False, sheet_name="Summary")
        vend_sum.to_excel(xw, index=False, sheet_name="ByVendor")
        needs_review.to_excel(xw, index=False, sheet_name="NeedsReview")

        # Post-format (autosize)
        wb = xw.book
        for name in ["Lines", "Summary", "ByVendor", "NeedsReview"]:
            ws = wb[name]
            autosize_columns(ws)

        # Add grand totals on Summary sheet (top row-ish)
        ws = wb["Summary"]
        # Compute totals (use dataframe to avoid Excel formulas)
        grand_due = float(pd.to_numeric(df["expense_due"], errors="coerce").fillna(0.0).sum())
        grand_paid = float(pd.to_numeric(df["expense_paid"], errors="coerce").fillna(0.0).sum())
        ws.insert_rows(1, amount=3)
        ws["A1"] = "Grand Totals"
        ws["A2"] = "Total Due"
        ws["B2"] = grand_due
        ws["A3"] = "Total Paid"
        ws["B3"] = grand_paid

    if debug:
        print(f"[debug] wrote xlsx: {out_xlsx}")


# -----------------------------
# CLI
# -----------------------------

def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--manifest", required=True, help="Path to manifest.jsonl")
    ap.add_argument("--text-dir", required=True, help="Directory containing <sha256>.txt OCR outputs")
    ap.add_argument("--out-xlsx", required=False, default="", help="Output xlsx path")
    ap.add_argument("--out-csv", required=False, default="", help="Output csv path")
    ap.add_argument("--only-canonical", action="store_true", help="Skip manifest rows where status==duplicate")
    ap.add_argument("--tax-year", type=int, default=2025, help="Tax year to prefer for date selection")
    ap.add_argument("--debug", action="store_true")
    args = ap.parse_args()

    manifest_path = Path(args.manifest)
    text_dir = Path(args.text_dir)

    if not manifest_path.exists():
        raise SystemExit(f"manifest not found: {manifest_path}")
    if not text_dir.exists():
        raise SystemExit(f"text-dir not found: {text_dir}")

    manifest_rows = read_manifest_jsonl(manifest_path)

    rows, missing_text = build_rows(
        manifest_rows=manifest_rows,
        text_dir=text_dir,
        tax_year=int(args.tax_year),
        only_canonical=bool(args.only_canonical),
        debug=bool(args.debug),
    )

    df = pd.DataFrame(rows)

    # Keep stable column order
    col_order = [
        "tax_year", "sha12", "sha256", "original_name", "manifest_status", "source_rel",
        "vendor", "category", "subcategory", "doc_status",
        "service_date", "statement_date", "due_date",
        "expense_due", "expense_paid",
        "likely_deductible", "notes", "snippet",
    ]
    df = df[[c for c in col_order if c in df.columns]]

    # Sort by service_date then vendor
    if "service_date" in df.columns:
        df = df.sort_values(["service_date", "vendor", "original_name"], kind="stable")

    out_csv = Path(args.out_csv) if args.out_csv else None
    out_xlsx = Path(args.out_xlsx) if args.out_xlsx else None

    if out_csv:
        write_csv(rows, out_csv)
        print(f"CSV:   {out_csv}")

    if out_xlsx:
        write_xlsx(df, out_xlsx, debug=bool(args.debug))
        print(f"Wrote: {out_xlsx}")

    if args.debug:
        print(f"[debug] manifest_rows={len(manifest_rows)} exported_rows={len(rows)} missing_text={missing_text}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
