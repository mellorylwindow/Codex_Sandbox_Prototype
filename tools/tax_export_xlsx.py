#!/usr/bin/env python3
"""
tax_export_xlsx.py — v2 (doc-level export, schema tolerant)

Reads:
- manifest.jsonl rows with:
    sha256, status, kind=receipt_asset, batch, original_name, dest_rel/src_rel
- parsed line json at:
    tax_intake/30_extracted/lines/<sha256>.json

Writes:
- Excel with 1 row per document, even if parsing is incomplete.

Goal:
- Never produce "Rows: 0" if there are documents.
"""

from __future__ import annotations

import argparse
import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def repo_root() -> Path:
    return Path(__file__).resolve().parents[1]


def utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def read_jsonl(path: Path) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    txt = path.read_text(encoding="utf-8", errors="replace")
    for line in txt.splitlines():
        if not line.strip():
            continue
        rows.append(json.loads(line))
    return rows


def safe_get(d: Dict[str, Any], key: str, default: Any = "") -> Any:
    v = d.get(key, default)
    return default if v is None else v


def first_date(dates: Any) -> str:
    if isinstance(dates, list) and dates:
        return str(dates[0])
    return ""


def join_amounts(amounts: Any, limit: int = 10) -> str:
    if not isinstance(amounts, list) or not amounts:
        return ""
    vals = []
    for a in amounts[:limit]:
        try:
            vals.append(f"{float(a):.2f}")
        except Exception:
            vals.append(str(a))
    if len(amounts) > limit:
        vals.append("…")
    return ", ".join(vals)


def autosize(ws, max_width: int = 60) -> None:
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        best = 0
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            best = max(best, len(str(v)))
        ws.column_dimensions[letter].width = min(max(10, best + 2), max_width)


def main(argv: Optional[List[str]] = None) -> int:
    ap = argparse.ArgumentParser(description="Export tax review spreadsheet (doc-level).")
    ap.add_argument("--manifest", required=True)
    ap.add_argument("--text-dir", default="tax_intake/30_extracted/text")
    ap.add_argument("--lines-dir", default="tax_intake/30_extracted/lines")
    ap.add_argument("--out", default="tax_intake/40_reports/tax_lines.xlsx")
    ap.add_argument("--only-canonical", action="store_true", help="Skip duplicates; include only ingested.")
    ap.add_argument("--debug", action="store_true")
    args = ap.parse_args(argv)

    repo = repo_root()

    man = Path(args.manifest)
    if not man.is_absolute():
        man = (repo / man).resolve()

    text_dir = (repo / args.text_dir).resolve()
    lines_dir = (repo / args.lines_dir).resolve()

    out_path = Path(args.out)
    if not out_path.is_absolute():
        out_path = (repo / out_path).resolve()
    out_path.parent.mkdir(parents=True, exist_ok=True)

    manifest_rows = read_jsonl(man)

    wb = Workbook()
    ws = wb.active
    ws.title = "Docs"

    headers = [
        "sha256",
        "status",
        "kind",
        "batch",
        "original_name",
        "date_guess",
        "vendor_guess",
        "total_guess",
        "amount_candidates",
        "amount_count",
        "ok",
        "error",
        "text_file",
        "lines_file",
        "src_rel",
        "dest_rel",
        "exported_at_utc",
    ]
    ws.append(headers)

    row_count = 0

    for r in manifest_rows:
        sha = str(safe_get(r, "sha256", "")).strip()
        if not sha:
            continue

        status = str(safe_get(r, "status", "")).strip().lower()
        if args.only_canonical and status == "duplicate":
            continue

        kind = str(safe_get(r, "kind", "")).strip()
        batch = str(safe_get(r, "batch", "")).strip()
        original_name = str(safe_get(r, "original_name", "")).strip()
        src_rel = str(safe_get(r, "src_rel", "")).strip()
        dest_rel = str(safe_get(r, "dest_rel", "")).strip()

        text_file = text_dir / f"{sha}.txt"
        lines_file = lines_dir / f"{sha}.json"

        # Defaults if line json missing
        ok = False
        err = "Missing lines JSON"
        vendor = ""
        date_guess = ""
        total_guess: Any = None
        amounts: Any = []

        if lines_file.exists():
            try:
                lj = json.loads(lines_file.read_text(encoding="utf-8", errors="replace"))
                ok = bool(lj.get("ok", False))
                err = lj.get("error") or ""
                vendor = str(lj.get("vendor", "") or lj.get("vendor_guess", "") or "").strip()
                date_guess = first_date(lj.get("dates"))
                total_guess = lj.get("total_guess", None)
                amounts = lj.get("amounts", []) or []
            except Exception as e:
                ok = False
                err = f"Failed reading lines JSON: {e!r}"

        ws.append([
            sha,
            status,
            kind,
            batch,
            original_name,
            date_guess,
            vendor,
            total_guess,
            join_amounts(amounts),
            (len(amounts) if isinstance(amounts, list) else 0),
            ok,
            err,
            str(text_file) if text_file.exists() else "",
            str(lines_file) if lines_file.exists() else "",
            src_rel,
            dest_rel,
            utc_now_iso(),
        ])
        row_count += 1

        if args.debug and row_count <= 3:
            print(f"[debug] export row {row_count}: {sha[:12]} vendor={vendor!r} date={date_guess!r} total={total_guess!r}")

    autosize(ws)

    # Simple README sheet for sanity
    ws2 = wb.create_sheet("Info")
    ws2.append(["generated_at_utc", utc_now_iso()])
    ws2.append(["manifest", str(man)])
    ws2.append(["text_dir", str(text_dir)])
    ws2.append(["lines_dir", str(lines_dir)])
    ws2.append(["rows_exported", row_count])
    ws2.append(["only_canonical", bool(args.only_canonical)])

    wb.save(out_path)

    print(f"Wrote: {out_path}")
    print(f"Rows:  {row_count}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
